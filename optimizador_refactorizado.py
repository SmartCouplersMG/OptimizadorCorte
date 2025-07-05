import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import openpyxl
import io

from pulp import *
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict, Counter
from openpyxl.drawing.image import Image as OpenpyxlImage
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as ReportLabImage
from reportlab.platypus import PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.utils import ImageReader

# Workaround para el problema de la librería reportlab
pt = inch / 72.0

def cargar_datos_entrada(ruta_archivo):
    """
    Carga y procesa los datos de entrada desde Excel.
    """
    try:
        df_inventario = pd.read_excel(ruta_archivo, sheet_name='Inventario')
        df_despieces = pd.read_excel(ruta_archivo, sheet_name='Despieces_Solicitados')
        df_parametros = pd.read_excel(ruta_archivo, sheet_name='Parametros')
        df_parametros.dropna(how='all', inplace=True)
        df_parametros.columns = df_parametros.columns.str.strip()
        df_parametros['Parametro'] = df_parametros['Parametro'].astype(str).str.strip()
        df_parametros['Valor'] = df_parametros['Valor'].astype(str).str.strip()
        print("✅ Archivo Excel y sus hojas leídos correctamente.")
        parametros = pd.Series(df_parametros['Valor'].values, index=df_parametros['Parametro']).to_dict()
        if 'UNIR_SOBRANTES' in parametros:
            parametros['UNIR_SOBRANTES'] = (str(parametros['UNIR_SOBRANTES']).strip().upper() == 'S')
        if 'DESPERDICIO_UTIL_MINIMO' in parametros:
            parametros['DESPERDICIO_UTIL_MINIMO'] = pd.to_numeric(parametros['DESPERDICIO_UTIL_MINIMO'], errors='coerce')
        return df_inventario, df_despieces, parametros
    except Exception as e:
        print(f"❌ ERROR al leer el archivo: {e}")
        return None, None, None

def procesar_datos_por_diametro(df_inventario, df_despieces):
    """
    Organiza y CLASIFICA los datos por diámetro, incluyendo todos los diámetros del inventario.
    """
    datos_procesados = {}
    df_inventario['Diametro'] = df_inventario['Diametro'].astype(str)
    df_despieces['Diametro'] = df_despieces['Diametro'].astype(str)
    for diametro in df_inventario['Diametro'].unique():
        inv_diametro = df_inventario[df_inventario['Diametro'] == diametro]
        des_diametro = df_despieces[df_despieces['Diametro'] == diametro]
        if not inv_diametro.empty:
            max_long_inventario = inv_diametro['Longitud'].max()
            despieces_dict = pd.Series(des_diametro.Cantidad.values, index=des_diametro.Longitud).to_dict()
            etiquetas_dict = pd.Series(des_diametro.Etiqueta.values, index=des_diametro.Longitud).to_dict()
            despieces_cortables, etiquetas_cortables, despieces_para_union, etiquetas_para_union = {}, {}, {}, {}
            for longitud, cantidad in despieces_dict.items():
                if longitud <= max_long_inventario:
                    despieces_cortables[longitud] = cantidad
                    etiquetas_cortables[longitud] = etiquetas_dict.get(longitud)
                else:
                    despieces_para_union[longitud] = cantidad
                    etiquetas_para_union[longitud] = etiquetas_dict.get(longitud)
            datos_procesados[diametro] = {
                'inventario': pd.Series(inv_diametro.Cantidad.values, index=inv_diametro.Longitud).to_dict(),
                'despieces': despieces_cortables, 'etiquetas_despieces': etiquetas_cortables,
                'despieces_requieren_union': despieces_para_union, 'etiquetas_requieren_union': etiquetas_para_union
            }
            print(f"🔩 Datos procesados para el diámetro: {diametro}")
    return datos_procesados

def generar_patrones(longitud_barra, longitudes_piezas):
    """
    Encuentra todas las combinaciones de corte (patrones) posibles.
    """
    patrones_validos = set()
    if not longitudes_piezas: return []
    longitudes_piezas.sort(reverse=True)
    min_pieza = min(longitudes_piezas)
    def encontrar_combos_recursivo(long_restante, patron_actual, start_index):
        if patron_actual: patrones_validos.add(tuple(sorted(patron_actual)))
        if long_restante < min_pieza: return
        for i in range(start_index, len(longitudes_piezas)):
            pieza = longitudes_piezas[i]
            if long_restante >= pieza:
                encontrar_combos_recursivo(long_restante - pieza, patron_actual + [pieza], i)
    encontrar_combos_recursivo(longitud_barra, [], 0)
    return [list(p) for p in patrones_validos]

def resolver_csp(datos_diametro, patrones_diametro):
    """
    Resuelve el CSP con un modelo avanzado que evita la sobreproducción
    y reporta faltantes. (Versión Final Corregida).
    """
    despieces = datos_diametro['despieces']
    plan_de_corte_final = {}
    piezas_faltantes = defaultdict(int)

    for long_barra, patrones in patrones_diametro.items():
        if not patrones or not despieces: continue

        problema = LpProblem(f"CSP_{long_barra}", LpMinimize)
        
        vars_patrones = {i: LpVariable(f"Patron_{i}", lowBound=0, cat='Integer') for i in range(len(patrones))}
        vars_excedente = {p: LpVariable(f"Excedente_{p}", lowBound=0, cat='Integer') for p in despieces.keys()}
        
        objetivo = lpSum(vars_patrones.values()) + 0.001 * lpSum(vars_excedente.values())
        problema += objetivo, "Objetivo_Minimizar_Barras_y_Excedente"

        # Restricción de igualdad para forzar la producción exacta + un excedente
        for pieza, cant_requerida in despieces.items():
            cantidad_producida = lpSum(patron.count(pieza) * vars_patrones[i] for i, patron in enumerate(patrones))
            problema += cantidad_producida == cant_requerida + vars_excedente[pieza], f"Req_{pieza}"
        
        problema.solve(pulp.PULP_CBC_CMD(timeLimit=60))
        
        plan_de_corte = {}
        if problema.status == 1:
            for i, var in vars_patrones.items():
                if value(var) > 0:
                    plan_de_corte[tuple(patrones[i])] = int(value(var))
        else:
            # Si no hay solución, todas las piezas son faltantes
            for pieza, cantidad in despieces.items():
                piezas_faltantes[pieza] += cantidad

        plan_de_corte_final[long_barra] = plan_de_corte
    
    # Calcular faltantes reales después de la optimización
    for pieza, cant_requerida in despieces.items():
        producido = sum(p.count(pieza) * repeticiones for plan in plan_de_corte_final.values() for p, repeticiones in plan.items())
        if producido < cant_requerida:
            piezas_faltantes[pieza] = cant_requerida - producido
            
    return plan_de_corte_final, dict(piezas_faltantes)

def _calcular_penalizacion(num_piezas):
    """
    Calcula un factor de penalización basado en el número de piezas usadas.
    """
    if num_piezas > 15: return 1.4
    if num_piezas > 9: return 1.3
    if num_piezas > 7: return 1.2
    if num_piezas > 5: return 1.1
    return 1.0

def encontrar_mejor_union_avanzado(longitud_objetivo, inventario_sobrantes):
    """
    Encuentra la mejor combinación de sobrantes usando un costo ponderado.
    """
    mejor_solucion = (float('inf'), None)
    piezas_disponibles = sorted([p for p, c in inventario_sobrantes.items() if c > 0], reverse=True)
    def buscar(long_actual, combo_actual, stock_restante, start_index):
        nonlocal mejor_solucion
        if long_actual >= longitud_objetivo:
            desperdicio = long_actual - longitud_objetivo
            penalizacion = _calcular_penalizacion(len(combo_actual))
            costo = desperdicio * penalizacion
            if costo < mejor_solucion[0]: mejor_solucion = (costo, combo_actual)
            return
        if start_index >= len(piezas_disponibles): return
        for i in range(start_index, len(piezas_disponibles)):
            pieza = piezas_disponibles[i]
            if stock_restante.get(pieza, 0) > 0:
                stock_restante[pieza] -= 1
                buscar(long_actual + pieza, combo_actual + [pieza], stock_restante, i)
                stock_restante[pieza] += 1
    buscar(0, [], inventario_sobrantes.copy(), 0)
    if mejor_solucion[1]:
        desperdicio_final = sum(mejor_solucion[1]) - longitud_objetivo
        return sorted(mejor_solucion[1]), desperdicio_final
    else:
        return None, None

def visualizar_plan_de_corte(diametro, plan_optimo, datos_diametro, params):
    """
    Crea una visualización estática del plan de corte.
    """
    fig, ax = plt.subplots(figsize=(12, 8)); y_pos = 0
    min_waste_param = params.get('DESPERDICIO_UTIL_MINIMO', float('inf'))
    if not datos_diametro['despieces']: return io.BytesIO()
    colores = plt.get_cmap('viridis', len(datos_diametro['despieces']))
    color_map = {pieza: colores(i) for i, pieza in enumerate(datos_diametro['despieces'])}
    for long_barra, plan in plan_optimo.items():
        for patron, repeticiones in plan.items():
            ax.text(-1, y_pos, f"Usar {repeticiones}x:", ha='right', va='center', fontsize=10)
            start_pos = 0
            for pieza in sorted(patron):
                rect = patches.Rectangle((start_pos, y_pos - 0.4), pieza, 0.8, facecolor=color_map.get(pieza, 'gray'), edgecolor='black')
                ax.add_patch(rect)
                ax.text(start_pos + pieza / 2, y_pos, str(pieza), ha='center', va='center', color='white', weight='bold')
                start_pos += pieza
            desperdicio = long_barra - sum(patron)
            if desperdicio > 0:
                if desperdicio >= min_waste_param:
                    rect = patches.Rectangle((start_pos, y_pos - 0.4), desperdicio, 0.8, facecolor='lightgray', edgecolor='black', hatch='//')
                else:
                    rect = patches.Rectangle((start_pos, y_pos - 0.4), desperdicio, 0.8, facecolor='white', edgecolor='black')
                ax.add_patch(rect)
                ax.text(start_pos + desperdicio / 2, y_pos, f"Sobra: {desperdicio}", ha='center', va='center', color='black', style='italic', fontsize=8)
            y_pos += 1
    ax.set_yticks([]); ax.set_xlabel("Longitud de la Barra (cm)")
    ax.set_title(f"Plan de Corte Visual para Diámetro {diametro}")
    if plan_optimo: ax.set_xlim(0, max(plan_optimo.keys()) * 1.1)
    ax.set_ylim(-1, y_pos); buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight'); plt.close(fig); buf.seek(0)
    return buf

def visualizar_plan_de_union(diametro, plan_de_union, inventarios_originales, params):
    """
    Genera una imagen del plan de unión, agrupando combinaciones idénticas.
    """
    num_combinaciones_unicas = sum(1 for combos in plan_de_union.values() for _ in combos)
    if num_combinaciones_unicas == 0:
        return io.BytesIO()

    fig_height = max(5, num_combinaciones_unicas * 1.5)
    fig, ax = plt.subplots(figsize=(12, fig_height))
    ax.set_title(f"Plan de Unión de Sobrantes - Diámetro {diametro}", fontsize=14)
    ax.set_xlabel("Longitud (cm)", fontsize=12)
    y_pos = 0
    max_len = 0

    for long_objetivo, combinaciones in sorted(plan_de_union.items()):
        # Dibuja la línea del objetivo una vez por cada grupo
        rect_objetivo = patches.Rectangle((0, y_pos + (len(combinaciones) - 1) * 1.5 / 2), long_objetivo, 0.4, fill=False, edgecolor='red', linewidth=1.5, linestyle='--')
        ax.add_patch(rect_objetivo)
        ax.text(long_objetivo / 2, y_pos + (len(combinaciones) - 1) * 1.5 + 0.5, f'Objetivo: {long_objetivo} cm', ha='center', va='center', fontsize=9, color='red')
        
        # Itera sobre cada combinación y sus repeticiones
        for combo, repeticiones in combinaciones.items():
            ax.text(-1, y_pos, f"Usar {repeticiones}x:", ha='right', va='center', fontsize=10)
            start_pos = 0
            for pieza in combo:
                es_original = pieza in inventarios_originales.get(str(diametro), {})
                color = 'slategray' if es_original else 'darkseagreen'
                rect = patches.Rectangle((start_pos, y_pos - 0.4), pieza, 0.8, color=color, edgecolor='black')
                ax.add_patch(rect)
                ax.text(start_pos + pieza / 2, y_pos, str(pieza), ha='center', va='center', color='white', weight='bold')
                start_pos += pieza
            
            max_len = max(max_len, start_pos)
            y_pos += 1.5

    ax.set_yticks([])
    ax.set_xlim(0, max_len * 1.1)
    ax.set_ylim(-1, y_pos)
    ax.grid(axis='x', linestyle=':', color='gray', alpha=0.6)
    plt.tight_layout()
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return buf

def generar_reporte_excel(resultados_finales, despieces_df, params):
    """
    Genera el reporte de Excel con todas las secciones finales y detalladas.
    """
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    sobrante_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    faltante_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    diferencia_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    min_waste_param = params.get('DESPERDICIO_UTIL_MINIMO', float('inf'))
    inventarios_originales = {diam: data['inventario_original'] for diam, data in resultados_finales.items() if isinstance(data, dict) and 'inventario_original' in data}

    for diametro, data in resultados_finales.items():
        if not isinstance(data, dict) or 'plan_optimo' not in data: continue
        ws = wb.create_sheet(f"Diametro_{diametro}"); current_row = 1; header_ranges = []
        
        piezas_para_union = data.get('despieces_requieren_union', {})
        if piezas_para_union:
            range_str = f'A{current_row}:C{current_row}'; ws[f'A{current_row}'] = "PIEZAS QUE EXCEDEN INVENTARIO (REQUIEREN UNIÓN)"; ws.merge_cells(range_str); header_ranges.append(range_str)
            current_row += 1; ws.append(["Longitud Requerida", "Cantidad", "Etiqueta"]); current_row += 1
            for longitud, cantidad in piezas_para_union.items():
                etiqueta = data['etiquetas_requieren_union'].get(longitud, 'N/A'); ws.append([longitud, cantidad, etiqueta]); current_row += 1
            current_row += 1
        
        start_row_resumen = current_row
        range_str = f'A{start_row_resumen}:B{start_row_resumen}'; ws[f'A{start_row_resumen}'] = "RESUMEN DE OPTIMIZACIÓN"; ws.merge_cells(range_str); header_ranges.append(range_str)
        current_row = start_row_resumen + 1
        metricas = data.get('metricas', {})
        ws.cell(row=current_row, column=1, value="Eficiencia del Material (Corte):"); ws.cell(row=current_row, column=2, value=metricas.get('eficiencia_corte', 0)).number_format = '0.00%'; current_row += 1
        ws.cell(row=current_row, column=1, value="Eficiencia del Material (Unión):"); ws.cell(row=current_row, column=2, value=metricas.get('eficiencia_union', 0)).number_format = '0.00%'; current_row += 1
        ws.cell(row=current_row, column=1, value="Eficiencia Total del Material:"); ws.cell(row=current_row, column=2, value=metricas.get('eficiencia_total', 0)).number_format = '0.00%'; current_row += 1
        ws.cell(row=current_row, column=1, value="Desperdicio Total por Corte (cm):"); ws.cell(row=current_row, column=2, value=metricas.get('desperdicio_corte', 0)); current_row += 1
        ws.cell(row=current_row, column=1, value="Desperdicio Útil por Corte (cm):"); ws.cell(row=current_row, column=2, value=metricas.get('desperdicio_util_corte', 0)); current_row += 1
        ws.cell(row=current_row, column=1, value="Desperdicio Útil Usado en Unión (cm):"); ws.cell(row=current_row, column=2, value=metricas.get('desperdicio_util_usado_union', 0)); current_row += 1
        ws.cell(row=current_row, column=1, value="Exceso Total por Unión (cm):"); ws.cell(row=current_row, column=2, value=metricas.get('exceso_union', 0)); current_row += 1
        ws.cell(row=current_row, column=1, value="Exceso Útil por Unión (cm):"); ws.cell(row=current_row, column=2, value=metricas.get('exceso_util_union', 0)); current_row += 1
        
        start_row_barras = current_row + 1
        range_str = f'A{start_row_barras}:B{start_row_barras}'; ws[f'A{start_row_barras}'] = "BARRAS DE INVENTARIO UTILIZADAS (CORTE)"; ws.merge_cells(range_str); header_ranges.append(range_str)
        current_row = start_row_barras + 1
        if data['barras_usadas_por_longitud']:
            ws.append(["Longitud de Barra", "Cantidad Usada"]); current_row += 1
            for long_barra, cantidad in data['barras_usadas_por_longitud'].items():
                ws.append([f"{long_barra} cm", cantidad]); current_row += 1
        else:
            ws.append(["No se utilizaron barras para corte."]); current_row += 1
        
        start_row_plan = current_row + 1
        range_str = f'A{start_row_plan}:E{start_row_plan}'; ws[f'A{start_row_plan}'] = "PLAN DE CORTE DETALLADO"; ws.merge_cells(range_str); header_ranges.append(range_str); current_row = start_row_plan + 1
        ws.append(["Barra de Inventario", "Patrón de Corte", "Repeticiones", "Desperdicio x Barra (cm)", "Tipo Desperdicio"]); current_row += 1
        for long_barra, plan in data['plan_optimo'].items():
            for patron, repeticiones in plan.items():
                desperdicio = long_barra - sum(patron)
                tipo_desperdicio = "Usable" if desperdicio >= min_waste_param else "No Usable"
                ws.append([long_barra, str(patron), repeticiones, desperdicio, tipo_desperdicio]); current_row += 1

        start_row_sobrantes_usados = current_row + 1
        range_str = f'A{start_row_sobrantes_usados}:C{start_row_sobrantes_usados}'; ws[f'A{start_row_sobrantes_usados}'] = "BARRAS UTILIZADAS EN UNIÓN"; ws.merge_cells(range_str); header_ranges.append(range_str); current_row = start_row_sobrantes_usados + 1
        if data.get('sobrantes_usados_en_union'):
            ws.append(["Fuente", "Longitud Sobrante", "Cantidad Usada"]); current_row += 1
            for longitud, cantidad in data['sobrantes_usados_en_union'].items():
                fuente = "Inventario Original" if longitud in inventarios_originales.get(diametro, {}) else "Desperdicio Útil"
                ws.append([fuente, longitud, cantidad]); current_row += 1
        else:
            ws.append(["No se usaron sobrantes para unión."]); current_row += 1
        
        start_row_union = current_row + 1
        range_str = f'A{start_row_union}:E{start_row_union}'; ws[f'A{start_row_union}'] = "PLAN DE UNIÓN DE SOBRANTES"; ws.merge_cells(range_str); header_ranges.append(range_str); current_row = start_row_union + 1
        if data.get('plan_de_union'):
            ws.append(["Pieza Objetivo", "Combinación de Sobrantes", "Repeticiones", "Exceso x Unión (cm)", "Tipo Exceso"]); current_row += 1
            for long_objetivo, combinaciones in data['plan_de_union'].items():
                for combo, repeticiones in combinaciones.items():
                    exceso = sum(combo) - long_objetivo
                    tipo_exceso = "Usable" if exceso >= min_waste_param else "No Usable"
                    ws.append([long_objetivo, str(combo), repeticiones, exceso, tipo_exceso]); current_row += 1
        else:
            ws.append(["No se realizaron uniones para este diámetro."]); current_row += 1
            
        start_row_faltantes = current_row + 1
        range_str = f'A{start_row_faltantes}:B{start_row_faltantes}'; ws[f'A{start_row_faltantes}'] = "PIEZAS FALTANTES"; ws.merge_cells(range_str); header_ranges.append(range_str); current_row = start_row_faltantes + 1
        piezas_faltantes_total = {**data.get('piezas_faltantes_corte',{}), **data.get('piezas_faltantes_union',{})}
        if piezas_faltantes_total:
            ws.append(["Longitud Pieza", "Cantidad Faltante"]); current_row += 1
            for longitud, cantidad in piezas_faltantes_total.items():
                 ws.append([longitud, cantidad])
                 for cell in ws[ws.max_row]: cell.fill = faltante_fill
        else:
            ws.append(["Se completaron todos los requerimientos."]); current_row += 1
        
        start_row_cumplimiento = current_row + 1
        range_str = f'A{start_row_cumplimiento}:C{start_row_cumplimiento}'; ws[f'A{start_row_cumplimiento}'] = "RESUMEN DE CUMPLIMIENTO DE PEDIDO"; ws.merge_cells(range_str); header_ranges.append(range_str); current_row = start_row_cumplimiento + 1
        ws.append(["Longitud", "Cantidad Solicitada", "Cantidad Obtenida"]); current_row += 1
        if data.get('resumen_cumplimiento'):
            for longitud, counts in data['resumen_cumplimiento'].items():
                solicitado, obtenido = counts['solicitado'], counts['obtenido']
                ws.append([longitud, solicitado, obtenido])
                if solicitado != obtenido:
                    for cell in ws[ws.max_row]: cell.fill = diferencia_fill
        
        for h_range in header_ranges:
            for row_tuple in ws[h_range]:
                for cell in row_tuple: cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        if data.get('grafico'):
            img_data = data['grafico']
            if isinstance(img_data, bytes) and img_data:
                img = OpenpyxlImage(io.BytesIO(img_data)); img.anchor = 'G2'; ws.add_image(img)
        if data.get('grafico_union'):
            img_data_union = data['grafico_union']
            if isinstance(img_data_union, bytes) and img_data_union:
                img_union = OpenpyxlImage(io.BytesIO(img_data_union)); img_union.anchor = 'G40'; ws.add_image(img_union)

    ws_inv = wb.create_sheet("Inventario Final")
    ws_inv['A1'] = "INVENTARIO FINAL CONSOLIDADO"; ws_inv.merge_cells('A1:D1')
    ws_inv.append(["Tipo de Material", "Diámetro", "Longitud", "Cantidad Restante"]); row = 3
    inventario_final = resultados_finales.get('inventario_final_consolidado', {})
    if inventario_final:
        for (diametro, longitud), cantidad in sorted(inventario_final.items(), key=lambda item: (float(item[0][0]), item[0][1])):
            if cantidad > 0:
                es_original = longitud in inventarios_originales.get(str(diametro), {})
                tipo = "Inventario Original" if es_original else "Sobrante Usable Generado"
                style_fill = None if es_original else sobrante_fill
                ws_inv.append([tipo, diametro, longitud, cantidad])
                if style_fill:
                    for col in range(1, 5): ws_inv.cell(row=row, column=col).fill = style_fill
                row += 1
    for cell_tuple in ws_inv["A1:D1"] + ws_inv["A2:D2"]:
        for cell in cell_tuple: cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    
    if not despieces_df.empty:
        ws_global = wb.create_sheet("Resumen Global por Etiqueta")
        ws_global['A1'] = "RESUMEN GLOBAL POR ETIQUETA DE ELEMENTO"; ws_global.merge_cells('A1:D1')
        ws_global.append(["Etiqueta", "Diámetro", "Longitud Requerida", "Cantidad Requerida"]);
        df_ordenado = despieces_df.sort_values(by=['Etiqueta', 'Diametro', 'Longitud'])
        for _, r in df_ordenado.iterrows():
            ws_global.append([r['Etiqueta'], r['Diametro'], r['Longitud'], r['Cantidad']])
        for cell_tuple in ws_global["A1:D1"] + ws_global["A2:D2"]:
            for cell in cell_tuple: cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)  # Rebobina el buffer al principio
    return output_buffer

def generar_pdf_resumen(resultados_finales, despieces_df, params, logo_path=None):
    """
    Genera un PDF con resumen ejecutivo, gráficos, tablas detalladas
    y una marca de agua en cada página.
    """
    # Canvas personalizado para marca de agua
    class WatermarkCanvas(pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_watermark()
                pdfcanvas.Canvas.showPage(self)
            pdfcanvas.Canvas.save(self)

        def draw_watermark(self):
            if logo_path:
                try:
                    logo = ImageReader(logo_path)
                    w, h = logo.getSize()
                    aspect = h / float(w)
                    display_w = 3.0 * inch
                    display_h = display_w * aspect
                    x = self._pagesize[0] - display_w - 0.5*inch
                    y = 0.5*inch
                    self.saveState()
                    self.setFillAlpha(0.2)
                    self.drawImage(logo, x, y, width=display_w, height=display_h, mask='auto')
                    self.restoreState()
                except Exception as e:
                    print(f"❌ WARN: no se pudo dibujar marca de agua ({e})")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            rightMargin=72, leftMargin=72,
                            topMargin=72, bottomMargin=72)
    styles = getSampleStyleSheet()
    story = []
    min_waste = params.get('DESPERDICIO_UTIL_MINIMO', float('inf'))
    inventarios_orig = {
        diam: data['inventario_original']
        for diam, data in resultados_finales.items()
        if isinstance(data, dict) and 'inventario_original' in data
    }

    # --- Cabecera ---
    story.append(Paragraph("<b>Resumen Ejecutivo - Optimización de Corte y Unión</b>", styles['h1']))
    story.append(Spacer(1, 12 * pt))
    story.append(Paragraph(
        "Este documento resume los resultados del proceso de optimización "
        "para el corte y unión de materiales, buscando minimizar el desperdicio "
        "y cumplir con los requerimientos.", styles['Normal']))
    story.append(Spacer(1, 12 * pt))
    story.append(Paragraph(
        "Metodología: Para determinar el corte óptimo se empleó un heurístico de "
        "programación dinámica basado en Tanir et al. (2016) en “One-dimensional Cutting Stock "
        "Problem with Divisible Items” :contentReference[oaicite:0]{index=0} y en el algoritmo BBP modificado de "
        "Berberler & Nuriyev [26].", styles['Normal']))
    story.append(Spacer(1, 24 * pt))

    # --- Detalle por diámetro ---
    primero = True
    for diam, data in resultados_finales.items():
        if not primero:
            story.append(PageBreak())
        primero = False

        if not isinstance(data, dict) or 'plan_optimo' not in data:
            continue
        story.append(Paragraph(f"<b>Diámetro: {diam}</b>", styles['h2']))
        story.append(Spacer(1, 6 * pt))

        # Métricas
        m = data.get('metricas', {})
        story.append(Paragraph(f"Eficiencia Corte: <b>{m.get('eficiencia_corte',0):.2%}</b>", styles['Normal']))
        story.append(Paragraph(f"Eficiencia Unión: <b>{m.get('eficiencia_union',0):.2%}</b>", styles['Normal']))
        story.append(Paragraph(f"Eficiencia Total: <b>{m.get('eficiencia_total',0):.2%}</b>", styles['Normal']))
        story.append(Spacer(1, 12 * pt))

        # Plan de corte
        story.append(Paragraph("<b>Plan de Corte Óptimo:</b>", styles['h3']))
        for L, plan in data['plan_optimo'].items():
            story.append(Paragraph(f"- Barra {L} cm:", styles['Normal']))
            for patron, repes in plan.items():
                waste = L - sum(patron)
                tip = "Usable" if waste >= min_waste else "No Usable"
                story.append(Paragraph(f"    • Patrón {patron} ×{repes} (Desp: {waste} cm – {tip})", styles['Normal']))
        story.append(Spacer(1, 12 * pt))

        # Gráfico de corte (si existe)
        if data.get('grafico'):
            imgbuf = data['grafico']
            img = ReportLabImage(io.BytesIO(imgbuf), width=6*inch)
            # escalar altura preservando proporción
            img.drawHeight = img.drawHeight * (6*inch / img.drawWidth)
            img.drawWidth = 6*inch
            # restringir al tamaño de la página
            pad = 6 * pt
            img._restrictSize(doc.width - 2*pad, doc.height - 2*pad)
            story.append(img)
            story.append(Spacer(1, 12 * pt))

        # Plan de unión
        if data.get('plan_de_union'):
            story.append(Paragraph("<b>Plan de Unión de Sobrantes:</b>", styles['h3']))
            for obj, combos in data['plan_de_union'].items():
                story.append(Paragraph(f"- Para {obj} cm:", styles['Normal']))
                for combo, repes in combos.items():
                    excess = sum(combo) - obj
                    tip_ex = "Usable" if excess >= min_waste else "No Usable"
                    story.append(Paragraph(f"    • Unión {combo} ×{repes} (Exceso: {excess} cm – {tip_ex})", styles['Normal']))
            story.append(Spacer(1, 12 * pt))

        # Gráfico de unión (si existe)
        if data.get('grafico_union'):
            imgbuf = data['grafico_union']
            img = ReportLabImage(io.BytesIO(imgbuf), width=6*inch)
            img.drawHeight = img.drawHeight * (6*inch / img.drawWidth)
            img.drawWidth = 6*inch
            pad = 6 * pt
            img._restrictSize(doc.width - 2*pad, doc.height - 2*pad)
            story.append(img)
            story.append(Spacer(1, 24 * pt))

    # --- Resumen global por etiqueta ---
    story.append(Paragraph("<b>Resumen Global por Etiqueta:</b>", styles['h2']))
    story.append(Spacer(1, 12 * pt))
    if not despieces_df.empty:
        df_ord = despieces_df.sort_values(by=['Etiqueta','Diametro','Longitud'])
        last = None
        for _, r in df_ord.iterrows():
            if r['Etiqueta'] != last:
                story.append(Spacer(1, 6*pt))
                story.append(Paragraph(f"<b>Etiqueta: {r['Etiqueta']}</b>", styles['h3']))
                last = r['Etiqueta']
            story.append(Paragraph(f"• {r['Cantidad']}×{r['Longitud']} cm (Ø{r['Diametro']})", styles['Normal']))
    else:
        story.append(Paragraph("No se encontraron elementos con etiquetas.", styles['Normal']))

    
    # ————————————— Inventario Final Consolidado —————————————
    story.append(PageBreak())
    story.append(Paragraph("<b>Inventario Final Consolidado</b>", styles['h2']))
    story.append(Spacer(1, 12 * pt))

    # Diccionario con (diametro, longitud) → cantidad restante
    inventario_final = resultados_finales.get('inventario_final_consolidado', {})

    if inventario_final:
        # Recorro ordenado por diámetro y longitud
        for (diametro, longitud), cantidad in sorted(
                inventario_final.items(),
                key=lambda item: (float(item[0][0]), item[0][1])
            ):
            # Determino si la pieza viene del inventario original o del sobrante usable
            es_original = longitud in resultados_finales[diametro]['inventario_original']
            tipo = "Inventario Original" if es_original else "Sobrante Usable Generado"
            # Agrego una línea por ítem
            story.append(Paragraph(
                f"- Tipo: {tipo} | Ø{diametro} | {longitud} cm | Cantidad: {cantidad}",
                styles['Normal']
            ))
    else:
        story.append(Paragraph("No hay inventario final consolidado.", styles['Normal']))

    # Construcción y guardado
    doc.build(story, canvasmaker=WatermarkCanvas)
    buf.seek(0) # Rebobina el buffer al principio
    return buf

def ejecutar_optimizacion(archivo_excel):
    """
    Función principal que orquesta todo el proceso de optimización,
    replicando la lógica completa del script original.
    """
    logs = []
    
    # --- FASE 1 y 2: CARGA Y PROCESAMIENTO ---
    logs.append("--- FASE 1: CARGA DE DATOS ---")
    inventario_df, despieces_df, params = cargar_datos_entrada(archivo_excel)
    if inventario_df is None:
        logs.append("Error al cargar los datos. Proceso detenido.")
        return None, None, None, logs

    min_waste_param = params.get('DESPERDICIO_UTIL_MINIMO', float('inf'))
    logs.append("\n--- FASE 2: PROCESAMIENTO Y ORGANIZACIÓN ---")
    datos_listos = procesar_datos_por_diametro(inventario_df, despieces_df)
    logs.append("FASE 2 COMPLETADA.")
    
    # --- INICIALIZACIÓN DE VARIABLES GLOBALES (LÓGICA RESTAURADA) ---
    resultados_finales = {}
    stock_sobrantes_global = defaultdict(int)
    for _, row in inventario_df.iterrows():
        stock_sobrantes_global[(str(row['Diametro']), row['Longitud'])] += row['Cantidad']
    
    piezas_globales_para_unir = {}
    
    # --- BUCLE PRINCIPAL (FASE DE CORTE) ---
    for diametro, datos in datos_listos.items():
        logs.append(f"\n--- PROCESANDO DIÁMETRO (CORTE): {diametro} ---")
        piezas_requeridas = list(datos.get('despieces', {}).keys())
        patrones_diametro = { long_barra: generar_patrones(long_barra, piezas_requeridas) for long_barra in datos.get('inventario', {}).keys() }
        
        # --- MANEJO DE LA TUPLA (COMO SOLICITASTE) ---
        plan_optimo, piezas_faltantes_corte = {}, {}
        if datos.get('despieces'):
            # Asumimos que resolver_csp devuelve la tupla (plan, faltantes)
            plan_optimo, piezas_faltantes_corte = resolver_csp(datos, patrones_diametro)

        # --- LÓGICA DE CÁLCULO DE LA FASE DE CORTE (RESTAURADA) ---
        barras_usadas_por_longitud = defaultdict(int)
        desperdicios_utiles_corte = defaultdict(int)
        for long_barra, plan in plan_optimo.items():
            for patron, repeticiones in plan.items():
                barras_usadas_por_longitud[long_barra] += repeticiones
                desperdicio_patron = long_barra - sum(patron)
                if desperdicio_patron >= min_waste_param:
                    desperdicios_utiles_corte[desperdicio_patron] += repeticiones
        
        for long_barra, cantidad_usada in barras_usadas_por_longitud.items():
            stock_sobrantes_global[(diametro, long_barra)] -= cantidad_usada
        for long, cant in desperdicios_utiles_corte.items():
            stock_sobrantes_global[(diametro, long)] += cant
        if datos.get('despieces_requieren_union'):
            piezas_globales_para_unir[diametro] = datos['despieces_requieren_union']

        # Almacenamos los resultados iniciales, incluyendo las piezas faltantes
        resultados_finales[diametro] = {
            'plan_optimo': plan_optimo, 'barras_usadas_por_longitud': dict(barras_usadas_por_longitud),
            'despieces_requieren_union': datos.get('despieces_requieren_union',{}),
            'etiquetas_requieren_union': datos.get('etiquetas_requieren_union',{}),
            'piezas_faltantes_corte': piezas_faltantes_corte, 'plan_de_union': {},
            'sobrantes_usados_en_union': {}, 'piezas_faltantes_union': {},
            'inventario_original': datos.get('inventario',{}), 'parametros': params
        }

    # --- FASE GLOBAL DE UNIÓN (LÓGICA RESTAURADA) ---
    inventario_provisional_union = stock_sobrantes_global.copy()
    if params.get('UNIR_SOBRANTES') and piezas_globales_para_unir:
        logs.append("\n--- FASE GLOBAL: Optimizando el plan de unión ---")
        inventarios_originales = {diam: data['inventario_original'] for diam, data in resultados_finales.items() if isinstance(data, dict)}
        for diametro, piezas_a_unir in piezas_globales_para_unir.items():
            plan_de_union_diametro, sobrantes_usados_en_union, piezas_faltantes_union = {}, defaultdict(int), defaultdict(int)
            inventario_union_diametro = {l:c for (d,l),c in inventario_provisional_union.items() if d == diametro}
            for long_objetivo, cant_requerida in sorted(piezas_a_unir.items(), reverse=True):
                uniones_encontradas = []
                for _ in range(cant_requerida):
                    # ¡ATENCIÓN! Esta función debe existir en tu script.
                    mejor_combo, exceso = encontrar_mejor_union_avanzado(long_objetivo, inventario_union_diametro)
                    if mejor_combo:
                        uniones_encontradas.append(tuple(sorted(mejor_combo)))
                        if exceso and exceso >= min_waste_param:
                            inventario_provisional_union[(diametro, exceso)] += 1
                        for pieza_usada in mejor_combo:
                            inventario_provisional_union[(diametro, pieza_usada)] -= 1
                            inventario_union_diametro[pieza_usada] -= 1
                            sobrantes_usados_en_union[pieza_usada] += 1
                    else:
                        piezas_faltantes_union[long_objetivo] += 1
                if uniones_encontradas:
                    plan_de_union_diametro[long_objetivo] = dict(Counter(uniones_encontradas))
            
            resultados_finales[diametro]['plan_de_union'] = plan_de_union_diametro
            resultados_finales[diametro]['sobrantes_usados_en_union'] = dict(sobrantes_usados_en_union)
            resultados_finales[diametro]['piezas_faltantes_union'] = dict(piezas_faltantes_union)
    
    resultados_finales['inventario_final_consolidado'] = {k: v for k, v in inventario_provisional_union.items() if v > 0}
    
    # --- FASE FINAL: CÁLCULO DE MÉTRICAS DETALLADAS Y REPORTES (LÓGICA RESTAURADA) ---
    logs.append("\n--- FASE FINAL: Calculando métricas y generando reportes ---")
    for diametro, data in resultados_finales.items():
        if not isinstance(data, dict) or 'plan_optimo' not in data:
            continue
        
        # --- BLOQUE DE CÁLCULO DE MÉTRICAS (RESTAURADO) ---
        metricas = {}
        longitud_total_piezas_cortadas = sum(sum(p) * r for plan in data['plan_optimo'].values() for p, r in plan.items())
        longitud_total_barras_corte = sum(l * c for l, c in data['barras_usadas_por_longitud'].items())
        metricas['eficiencia_corte'] = longitud_total_piezas_cortadas / longitud_total_barras_corte if longitud_total_barras_corte > 0 else 0
        metricas['desperdicio_corte'] = longitud_total_barras_corte - longitud_total_piezas_cortadas
        longitud_total_piezas_unidas = sum(l * sum(c.values()) for l, c in data.get('plan_de_union', {}).items())
        sobrantes_usados = data.get('sobrantes_usados_en_union', {})
        longitud_sobrantes_usados_original = sum(l * c for l, c in sobrantes_usados.items() if l in data.get('inventario_original', {}))
        longitud_sobrantes_usados_desperdicio = sum(l * c for l, c in sobrantes_usados.items() if l not in data.get('inventario_original', {}))
        longitud_total_material_nuevo_union = longitud_sobrantes_usados_original
        metricas['eficiencia_union'] = longitud_total_piezas_unidas / longitud_total_material_nuevo_union if longitud_total_material_nuevo_union > 0 else 1.0
        longitud_total_producida = longitud_total_piezas_cortadas + longitud_total_piezas_unidas
        longitud_total_consumida_nueva = longitud_total_barras_corte + longitud_sobrantes_usados_original
        metricas['eficiencia_total'] = longitud_total_producida / longitud_total_consumida_nueva if longitud_total_consumida_nueva > 0 else 0
        # --- INSERTA BLOQUE ---
        desperdicios_corte_dict = defaultdict(int)
        for long_barra, plan in data['plan_optimo'].items():
            for patron, repeticiones in plan.items():
                desperdicio = long_barra - sum(patron)
                if desperdicio >= min_waste_param:
                    desperdicios_corte_dict[desperdicio] += repeticiones
        metricas['desperdicio_util_corte'] = sum(l*c for l,c in desperdicios_corte_dict.items())
        metricas['desperdicio_util_usado_union'] = longitud_sobrantes_usados_desperdicio

        excesos_union_dict = defaultdict(int)
        for obj, combos in data.get('plan_de_union', {}).items():
            for combo, reps in combos.items():
                exceso = sum(combo) - obj
                excesos_union_dict[exceso] += reps
        metricas['exceso_union'] = sum(l*c for l,c in excesos_union_dict.items())
        metricas['exceso_util_union'] = sum(l*c for l,c in excesos_union_dict.items() if l >= min_waste_param)
        # --- FIN DEL BLOQUE A INSERTAR ---
        
        data['metricas'] = metricas

        # --- BLOQUE DE CÁLCULO DE CUMPLIMIENTO (RESTAURADO) ---
        cortes_totales_por_pieza = defaultdict(int)
        for plan in data['plan_optimo'].values():
            for patron, repeticiones in plan.items():
                for pieza in patron:
                    cortes_totales_por_pieza[pieza] += repeticiones
        
        solicitado_total = {**datos_listos[diametro].get('despieces', {}), **datos_listos[diametro].get('despieces_requieren_union', {})}
        uniones_totales_por_pieza = defaultdict(int)
        for long_objetivo, combinaciones in data.get('plan_de_union', {}).items():
            uniones_totales_por_pieza[long_objetivo] = sum(combinaciones.values())
            
        resumen_cumplimiento = {}
        for longitud, cant_solicitada in solicitado_total.items():
            obtenido = cortes_totales_por_pieza.get(longitud, 0) + uniones_totales_por_pieza.get(longitud, 0)
            resumen_cumplimiento[longitud] = {'solicitado': cant_solicitada, 'obtenido': obtenido}
        data['resumen_cumplimiento'] = resumen_cumplimiento
        
        # --- Generación de gráficos (ya debería estar presente) ---
        data['grafico'] = visualizar_plan_de_corte(diametro, data['plan_optimo'], datos_listos[diametro], params).getvalue()
        if data.get('plan_de_union'):
            data['grafico_union'] = visualizar_plan_de_union(diametro, data['plan_de_union'], inventarios_originales, params).getvalue()

    # Generación de archivos de salida en memoria
    archivo_excel_resultado = generar_reporte_excel(resultados_finales, despieces_df, params)
    
    # ¡ATENCIÓN! Esta función debe existir y devolver un buffer en memoria.
    archivo_pdf_resultado = generar_pdf_resumen(resultados_finales, despieces_df, params, logo_path='LOGO.jpg')
    
    logs.append("\n--- PROCESO FINALIZADO ---")
    
    # La función ahora devuelve los dos archivos, los resultados y los logs
    return archivo_excel_resultado, archivo_pdf_resultado, resultados_finales, logs